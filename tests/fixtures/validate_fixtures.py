#!/usr/bin/env python3
"""
Validate test fixtures.

Ensures all fixtures are valid and meet requirements.
"""

import wave
from pathlib import Path


def validate_pdf(path: Path) -> dict:
    """Validate PDF fixture."""
    try:
        from pypdf import PdfReader

        pdf = PdfReader(path)
        return {
            "valid": True,
            "size": path.stat().st_size,
            "details": f"{len(pdf.pages)} pages, {path.stat().st_size} bytes",
        }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def validate_docx(path: Path) -> dict:
    """Validate DOCX fixture."""
    try:
        from docx import Document

        doc = Document(path)
        return {
            "valid": True,
            "size": path.stat().st_size,
            "details": f"{len(doc.paragraphs)} paragraphs, {path.stat().st_size} bytes",
        }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def validate_pptx(path: Path) -> dict:
    """Validate PPTX fixture."""
    try:
        from pptx import Presentation

        prs = Presentation(path)
        return {
            "valid": True,
            "size": path.stat().st_size,
            "details": f"{len(prs.slides)} slides, {path.stat().st_size} bytes",
        }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def validate_xlsx(path: Path) -> dict:
    """Validate XLSX fixture."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb.active
        rows = sum(1 for _ in ws.iter_rows())
        return {
            "valid": True,
            "size": path.stat().st_size,
            "details": f"{rows} rows, {path.stat().st_size} bytes",
        }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def validate_text(path: Path) -> dict:
    """Validate a text-based fixture (txt, html, md, csv, adoc, vtt)."""
    try:
        content = path.read_text()
        lines = len(content.splitlines())
        return {
            "valid": len(content) > 0,
            "size": path.stat().st_size,
            "details": f"{lines} lines, {path.stat().st_size} bytes",
        }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def validate_image(path: Path) -> dict:
    """Validate an image fixture."""
    try:
        from PIL import Image

        img = Image.open(path)
        img.verify()
        return {
            "valid": True,
            "size": path.stat().st_size,
            "details": f"{img.size[0]}x{img.size[1]} {img.format}, {path.stat().st_size} bytes",
        }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def validate_mp3(path: Path) -> dict:
    """Validate MP3 fixture."""
    try:
        data = path.read_bytes()
        has_id3 = data[:3] == b"ID3"
        has_sync = b"\xff\xfb" in data or b"\xff\xfa" in data
        return {
            "valid": has_id3 or has_sync,
            "size": path.stat().st_size,
            "details": f"{'ID3 tag' if has_id3 else 'No ID3'}, {path.stat().st_size} bytes",
        }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def validate_wav(path: Path) -> dict:
    """Validate WAV fixture."""
    try:
        with wave.open(str(path), "r") as wav:
            duration = wav.getnframes() / wav.getframerate()
            return {
                "valid": True,
                "size": path.stat().st_size,
                "details": f"{wav.getnchannels()}ch, {wav.getframerate()}Hz, {duration:.1f}s, {path.stat().st_size} bytes",
            }
    except Exception as e:
        return {"valid": False, "error": str(e)}


def main():
    """Validate all fixtures."""
    fixtures_dir = Path(__file__).parent

    print("Validating test fixtures...\n")

    validators = {
        # Documents
        "sample.pdf": validate_pdf,
        "sample.docx": validate_docx,
        "sample.pptx": validate_pptx,
        "sample.xlsx": validate_xlsx,
        "sample.txt": validate_text,
        "sample.html": validate_text,
        "sample.md": validate_text,
        "sample.csv": validate_text,
        "sample.adoc": validate_text,
        "sample.vtt": validate_text,
        # Images
        "sample.png": validate_image,
        "sample.jpg": validate_image,
        "sample.tiff": validate_image,
        "sample.bmp": validate_image,
        "sample.webp": validate_image,
        # Audio
        "sample.mp3": validate_mp3,
        "sample.wav": validate_wav,
    }

    all_valid = True
    total_size = 0

    for filename, validator in validators.items():
        path = fixtures_dir / filename
        if not path.exists():
            print(f"  ✗ {filename}: NOT FOUND")
            all_valid = False
            continue

        result = validator(path)
        if result["valid"]:
            print(f"  ✓ {filename}: {result['details']}")
            total_size += result["size"]
        else:
            print(f"  ✗ {filename}: INVALID - {result.get('error', 'Unknown error')}")
            all_valid = False

    print(f"\n{'='*60}")
    print(f"Total size: {total_size:,} bytes ({total_size / 1024:.1f} KB)")
    print(f"Status: {'✓ All valid' if all_valid else '✗ Some fixtures invalid'}")

    return 0 if all_valid else 1


if __name__ == "__main__":
    exit(main())
